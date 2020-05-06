import yaml
from openpyxl import load_workbook, Workbook


excel_source = load_workbook('../source.xlsx')
cols = [0,2,1,3]
for shit in excel_source.sheetnames:
    vals = []
    sheet = excel_source.get_sheet_by_name(shit)
    for row in sheet.iter_rows():
        cells = []
        for idx, cell in enumerate(row):
            if idx in cols:
                if cell.value is None:
                    cell.value = f'Missing value in fucking motherfucker {shit}'
            cells.append(cell.value)
        vals.append(cells)

    hate_you = []
    for d in vals[1:]:

        if d[2] == '@':
            name_value = ''
        else:
            name_value = d[2]

        dont_eat_cheese_its_yuky = yaml.MappingNode(
            tag='tag:yaml.org,2002:map',
            value=[
                (yaml.ScalarNode(tag='tag:yaml.org,2002:str', value=name_value),
                 yaml.MappingNode(tag='tag:yaml.org,2002:map', value=[
                    (yaml.ScalarNode(tag='tag:yaml.org,2002:str', value='type'),
                     yaml.ScalarNode(tag='tag:yaml.org,2002:str', value=d[1])),
                    (yaml.ScalarNode(tag='tag:yaml.org,2002:str', value='value'),
                     yaml.ScalarNode(tag='tag:yaml.org,2002:str', value=d[3]))
                 ])
                 )
                 ])
        hate_you.append(yaml.safe_load(yaml.serialize(dont_eat_cheese_its_yuky)))

    with open(shit + '.yaml', 'w+') as file:
        documents = yaml.dump(hate_you, file, explicit_start=True)
